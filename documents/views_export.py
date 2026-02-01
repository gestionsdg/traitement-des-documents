# documents/views_export.py
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import (
    DocumentEntrant,
    Chrono, Gouvernement, Decision,
    PaiementPrestationsSociales, AllocationsFamiliales, AllocationsPrenatales,
    OrdreMission
)


def _autosize_columns(ws):
    """Ajuste la largeur des colonnes selon le contenu."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def _fmt_date(d):
    """
    ✅ Format : jour mois année (ex: 06 01 2026)
    (Affichage stable dans Excel car c'est du texte)
    """
    if not d:
        return ""
    try:
        return d.strftime("%d %m %Y")
    except Exception:
        return str(d)


def export_entrants_excel(request):
    """
    Exporte DocumentEntrant vers Excel.
    Colonnes :
    N° Ord (auto), N° entrée, Date, Expéditeur, Objet, Prov., Dest., Annotations,
    Date retour, Objet retour

    ✅ Corrections :
    - N° Ord auto 1..N (recommence à 1)
    - Tri croissant par Date (puis N° entrée)
    - Dates au format : jour mois année (dd mm YYYY)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Entrants"

    headers = [
        "N° Ord", "N° entrée", "Date", "Expéditeur", "Objet",
        "Prov.", "Dest.", "Annotations", "Date retour", "Objet retour"
    ]
    ws.append(headers)

    # ✅ Tri croissant : Date puis N° entrée (et pk pour stabilité)
    qs = DocumentEntrant.objects.all().order_by("date_reception", "numero_entree", "pk")

    # ✅ N° Ord auto 1..N
    for i, d in enumerate(qs, start=1):
        ws.append([
            i,
            d.numero_entree,
            _fmt_date(d.date_reception),
            d.expediteur,
            d.objet,
            d.nature_document,  # KIN / PROV
            d.destinataire,
            d.annotations_autorite,
            _fmt_date(d.date_document_retourne),
            d.objet_document_retourne or "",
        ])

    _autosize_columns(ws)

    now = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M")
    filename = f"documents_entrants_{now}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _export_sortants_to_sheet(ws, queryset, label):
    """
    Remplit une feuille Excel pour un type de sortant.
    Colonnes : N° Ord(auto), N° registre, Date sortie, Objet, Destinataire, Émetteur, Statut

    ✅ Corrections :
    - N° Ord auto 1..N (recommence à 1 pour CHAQUE feuille)
    - Dates au format : dd mm YYYY
    """
    ws.append([label])
    ws.append([])

    headers = ["N° Ord", "N° registre", "Date sortie", "Objet", "Destinataire", "Émetteur", "Statut"]
    ws.append(headers)

    for i, d in enumerate(queryset, start=1):
        ws.append([
            i,  # ✅ N° Ord auto
            d.numero_registre,
            _fmt_date(d.date_sortie),
            d.objet,
            d.destinataire,
            d.emetteur,
            d.statut,
        ])

    ws.append([])


def _export_ordre_mission_to_sheet(ws, queryset, label):
    """
    ✅ Export spécifique OrdreMission.
    Colonnes :
    N° Ord(auto), N° registre, Date sortie, Objet, Personne(s) désigné(es),
    Statut, Destination mission, Nombre de jours

    ✅ Corrections :
    - N° Ord auto 1..N (recommence à 1)
    - Dates au format : dd mm YYYY
    """
    ws.append([label])
    ws.append([])

    headers = [
        "N° Ord", "N° registre", "Date sortie", "Objet",
        "Personne(s) désigné(es)", "Statut", "Destination mission", "Nombre de jours"
    ]
    ws.append(headers)

    for i, d in enumerate(queryset, start=1):
        ws.append([
            i,  # ✅ N° Ord auto
            d.numero_registre,
            _fmt_date(d.date_sortie),
            d.objet,
            d.destinataire,  # destinataire = personne(s) désigné(es)
            d.statut,
            d.destination_mission or "",
            d.nombre_jours if d.nombre_jours is not None else "",
        ])

    ws.append([])


def export_sortants_excel(request):
    """
    Exporte les Sortants vers un seul fichier Excel (plusieurs onglets).

    ✅ Corrections :
    - Tri croissant par Date sortie (puis N° registre)
    - N° Ord auto 1..N dans chaque onglet
    - Dates au format : dd mm YYYY
    """
    wb = Workbook()
    wb.remove(wb.active)

    sortants_classiques = [
        ("Chrono", Chrono.objects.all()),
        ("Gouvernement", Gouvernement.objects.all()),
        ("Décision", Decision.objects.all()),
        ("Paiement Prestations Sociales", PaiementPrestationsSociales.objects.all()),
        ("Allocations familiales", AllocationsFamiliales.objects.all()),
        ("Allocations prénatales", AllocationsPrenatales.objects.all()),
    ]

    # 1) Onglets classiques
    for sheet_name, qs in sortants_classiques:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limite 31 caractères
        # ✅ Tri croissant : date_sortie puis numero_registre (et pk si dispo)
        qs = qs.order_by("date_sortie", "numero_registre", "pk")
        _export_sortants_to_sheet(ws, qs, sheet_name)
        _autosize_columns(ws)

    # 2) Onglet Ordre de mission (spécifique)
    ws = wb.create_sheet(title="Ordre de mission")
    qs = OrdreMission.objects.all().order_by("date_sortie", "numero_registre", "pk")
    _export_ordre_mission_to_sheet(ws, qs, "Ordre de mission")
    _autosize_columns(ws)

    now = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M")
    filename = f"documents_sortants_{now}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
